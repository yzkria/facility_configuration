# importing openpyxl module
import openpyxl as xl;
import string;
from openpyxl import load_workbook;
from copy import copy;
#from xlwt import Workbook


style_attrs = ["alignment", "border", "fill", "font", "number_format", "protection"]
# opening the source excel file
filename ="D:\\source.xlsx"
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]

# create the destination excel file
#filename1 ="D:\\target.xlsx"
wb2 = xl.Workbook()
ws2 = wb2.worksheets[0]

# create excel sheet for wrong rows

wb3 = xl.Workbook()
ws3 = wb3.worksheets[0]

# calculate total number of rows and
# columns in source excel file
mr = ws1.max_row
mc = ws1.max_column

delete_rows = 0
# copy source attributes
def copy_attrs(src, dst, attrs=style_attrs):
    """Copy attributes from src to dst. Attributes are shallow-copied to avoid
    TypeError: unhashable type: 'StyleProxy'"""
    for name in attrs:
        setattr(dst, name, copy(getattr(src, name)))


def copy_column_attrs(worksheet_src, worksheet_dst, attrs=style_attrs + ["width"]):
    """Copy ColumnDimension properties from worksheet_src to worksheet_dst.
    Only properties listed in attrs will be copied."""
    for column, dimensions in worksheet_src.column_dimensions.items():
        copy_attrs(
            src=dimensions,
            dst=worksheet_dst.column_dimensions[column],
            attrs=attrs,
        )


def copy_row_attrs(worksheet_src, worksheet_dst, attrs=style_attrs + ["height"]):
    """Copy RowDimension properties from worksheet_src to worksheet_dst.
    Only properties listed in attrs will be copied."""
    for row, dimensions in worksheet_src.row_dimensions.items():
        copy_attrs(
            src=dimensions,
            dst=worksheet_dst.row_dimensions[row],
            attrs=style_attrs + ["height"],
        )

#copy header to ready and incorrect excel sheets
for x in range(1,mc+1):
    c = ws1.cell(row = 1,column = x)
    ws2.cell(row = 1, column = x).value = c.value
    ws3.cell(row = 1, column = x).value = c.value


# copying the cell values from source to ready excel file
for i in range (2, mr + 1):
    for j in range (1, mc + 1):
        # reading cell value from source excel file
        cell = ws1.cell(row = i, column = j)
        #Check which column we are now
        col_header = ws1.cell(row = 1, column = j)
        if col_header.value == "NATIONAL_ID_NUM":
            if cell.value =="":
                # write all row to incorrect sheet and skip rest of columns at same row
                #ws3.cell(row = i, column = j).value = cell.value
                for y in range(1,mc+1):
                    cc = ws1.cell(row = i,column = y)
                    ws3.cell(row = i, column = y).value = cc.value
                    delete_rows = i
            elif len(str(cell.value))==14:
                # writing the read value to ready excel file
                ws2.cell(row = i, column = j).value = cell.value
                ws2.cell(row = i, column = j).number_format = '0'
            else:
                # write all row to incorrect sheet and skip rest of columns at same row
                #ws3.cell(row = i, column = j).value = cell.value
                for y in range(1,mc+1):
                    cc = ws1.cell(row = i,column = y)
                    ws3.cell(row = i, column = y).value = cc.value
                    delete_rows = i
        elif col_header.value == "PRACTITIONER_NAME_ENGLISH":
            #remove special characters
            result = ""
            for ch in cell.value:
                # If char is not punctuation, add it to the result.
                if ch not in string.punctuation:
                    result += ch
                    # writing the read value to destination excel file
                    ws2.cell(row = i, column = j).value = result
        elif col_header.value == "SHORT_NAME":
            #remove special characters
            result = ""
            for ch in cell.value:
                # If char is not punctuation, add it to the result.
                if ch not in string.punctuation:
                    result += ch
                    # writing the read value to destination excel file
                    ws2.cell(row = i, column = j).value = result
        elif col_header.value == "DATE_OF_BIRTH":
            # writing the read value to ready excel file
            ws2.cell(row = i, column = j).value = cell.value
            ws2.cell(row = i, column = j).number_format = 'dd/mm/yyyy'
        else:
            # writing the read value to ready excel file
            ws2.cell(row = i, column = j).value = cell.value
if delete_rows>1:
    ws2.delete_rows(delete_rows)
#print(delete_rows-1)
# copy attripute from source sheet to ready sheet
#copy_attrs(ws1,ws2)
copy_column_attrs(ws1, ws2)
copy_row_attrs(ws1, ws2)

# saving the destination excel file
wb2.save("D:\\ready.xlsx")

# copy attripute from source sheet to incorrect sheet
#copy_attrs(ws1,ws3)
copy_column_attrs(ws1, ws3)
copy_row_attrs(ws1, ws3)
# save incorrect names in seperate sheet
wb3.save("D:\\incorrect.xlsx")

