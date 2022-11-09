# importing openpyxl module
import openpyxl as xl;
import pandas as pd
import numpy as np
import string;
from openpyxl import load_workbook;
from copy import copy;
import win32com.client;
from tkinter import Tk     # from tkinter import Tk for Python 3.x
from tkinter.filedialog import askopenfilename ,asksaveasfilename
#from xlwt import Workbook

headers_dic={}
p_spec = {'GENERAL SURGERY':'GENS',
'GENERAL SURGERY SPLTY':'GENS',
'INTERNAL MEDICINE SPLTY':'MED',
'CARDIOLOGY':'CARS',
'PEDIATRIC SPLTY':'PE',
'ICU':'ICUS',
'EMERGENCY MED-SPLTY':'EMMS',
'ER':'ER',
'FAMILY PLANNING':'FAMP',
'NEPHROLOGY':'NEPS',
'ORTHOPEDICS SPLTY':'ORPS',
'ANESTHESIOLOGY':'ANES',
'ANESTHESIA':'ANES',
'NURSE':'NS',
'PHARMACY':'PH',
'NEONATOLOGY':'NEOS',
'DENTISTRY SPLTY':'DN',
'GENERAL OPTHOMOLOGY SPLY':'OPTS',
'FAMILY MEDICINE SPLTY':'FAMS',
'FAMILY MEDICINE':'FAMS',
'NEUROLOGY':'NEPY',
'HEAMATOLOGY':'HMT',
'PHYSIOTHERAPY SPL':'PHYS',
'PULMONOLOGY':'PU',
'E.N.T':'ENTS',
'RADIOLOGY':'RAD',
'VASCULAR SURGERY SPECIALITY':'VSS',
'GASTROENTEROLOG':'GE',
'DERMATOLOGY':'DRMS',
'PSYCHIATRIST SPEC':'PSYS',
'PLASTIC SURGERY':'PLAS',
'CARDIOTHORACIC SPL':'CTSP',
'PEDIATRIC CARDIOLOGY':'PAC',
'NEUROSURGERY':'NESU',
'UROLOGY':'UROS',
'SURGICAL ORTHOPEDIC':'SO',
'LABORATORY':'LA',
'PEDIATRIC SURGERY':'PASU',
'ORTHOPEDIC SURGERY':'OS',
'MAXILLOFACIAL':'MOF',
'GLUCOMA OPHTHALMOLOGY':'GLOP',
'CARDIOLOGY VASCULAR':'CR',
'OBSTETRICS & GYNEACOLOGY':'OB',
'CLINICAL PATHOLOGY':'PATH',
'INTENSIVE CARE SPLTY':'INCS',
'DIALYSIS SPL':'DISP',
'CLINICAL PATHOLOGY SPLTY':'CLPA'}
# opening the source excel file
# Tk().withdraw() # we don't want a full GUI, so keep the root window from appearing
# source_file = askopenfilename(title="Choose Sheet You Need To Validate",filetypes=[("Excel Workbook", ".xlsx")],defaultextension=".xlsx") # show an "Open" dialog box and return the path to the selected file
source_file ="D:\\new.xlsx"
# if source_file:
wb1 = xl.load_workbook(source_file)
ws1 = wb1.worksheets[0]
# else:
    # Print(" You Must Select Source File to Continue..")

# create ready excel file
wb2 = xl.Workbook()
ws2 = wb2.worksheets[0]

# create excel sheet for wrong rows
wb3 = xl.Workbook()
ws3 = wb3.worksheets[0]

# create middle excel sheet
wb4 = xl.Workbook()
ws4 = wb4.worksheets[0]

# define list to have all index of rows that need tobe deleted from ready sheet
del_rows = []
# function to remove empty rows
# def remove(sheet):
#     # iterate the sheet by rows
#     for row in sheet.iter_rows():
#         # all() return False if all of the row value is None
#         if not all(cell.value for cell in row):
#             # detele the empty row
#             sheet.delete_rows(row[0].row, 1)
#             # recursively call the remove() with modified sheet data
#             remove(sheet)
#             return

#copy header to ready ,middle and incorrect excel sheets
for x in range(1,ws1.max_column+1):
    c = ws1.cell(row = 1,column = x)
    result = ""
    for ch in str(c.value):
        # If char is not punctuation, add it to the result.
        if ch not in string.punctuation:
            result += ch
    ws2.cell(row = 1, column = x).value = result.upper().strip().replace(" ","")
    ws3.cell(row = 1, column = x).value = result.upper().strip().replace(" ","")
    ws4.cell(row = 1, column = x).value = result.upper().strip().replace(" ","")
    headers_dic[result.upper().strip().replace(" ","")] = x

#add two extra columns PRACTITIONER_CONSULTANT ,PRACTITIONER_SPECIALIST to three sheets
ws2.cell(row=1, column=ws2.max_column+1).value = "PRACTITIONER_CONSULTANT"
ws2.cell(row=1, column=ws2.max_column+1).value = "PRACTITIONER_SPECIALIST"
ws3.cell(row=1, column=ws3.max_column+1).value = "PRACTITIONER_CONSULTANT"
ws3.cell(row=1, column=ws3.max_column+1).value = "PRACTITIONER_SPECIALIST"
ws4.cell(row=1, column=ws4.max_column+1).value = "PRACTITIONER_CONSULTANT"
ws4.cell(row=1, column=ws4.max_column+1).value = "PRACTITIONER_SPECIALIST"
headers_dic["PRACTITIONER_CONSULTANT"] = ws4.max_column-1
headers_dic["PRACTITIONER_SPECIALIST"] = ws4.max_column

# copying the cell values from source to middle excel file
for r in range (2, ws1.max_row + 1):
    for cl in range (1, ws1.max_column + 1):
        # reading cell value from source excel file
        cell = ws1.cell(row = r, column = cl)

        # writing the read value to ready excel file
        ws4.cell(row = r, column = cl).value = cell.value

# add short name in english as copy from long name to middle excel sheet
for n in range(2, ws4.max_row+1):
    #copy from wb1
    col = ws4.cell(row=n, column=headers_dic['PRACTITIONERNAMEENGLISH'])
    #paste in ws2
    ws4.cell(row=n, column=headers_dic['SHORTNAME'], value=col.value)

# add birth date from national id number to middle excel sheet
for d in range(2, ws4.max_row+1):
    #copy from wb1
    coll = ws4.cell(row=d, column=headers_dic['NATIONALIDNUM'])
    if (coll.value != None) and (len(str(coll.value))==14):
        #paste in ws2
        res = []
        for m in str(coll.value):
            res.append(int(m))
        birth_date = str(res[3]) + str(res[4])+'/'+str(res[5]) + str(res[6])+'/'+'19'+str(res[1]) + str(res[2])
        ws4.cell(row=d, column=headers_dic['DATEOFBIRTH'], value = birth_date)

# add practitioner consultant and specialist falgs based on position to middle excel sheet
for p in range(2, ws4.max_row+1):
    #copy from wb1
    coll = ws4.cell(row=p, column=headers_dic['POSITION'])
    if (coll.value == "Consultant"):
        ws4.cell(row=p, column=headers_dic['PRACTITIONER_CONSULTANT'], value = "Y")
        ws4.cell(row=p, column=headers_dic['PRACTITIONER_SPECIALIST'], value = "N")
    else:
        ws4.cell(row=p, column=headers_dic['PRACTITIONER_SPECIALIST'], value = "Y")

# add primary speciality code from primary speciality name to middle excel sheet
for s in range(2, ws4.max_row+1):
#     #copy from wb1
    col = ws4.cell(row=s, column=headers_dic['PRIMARYSPECIALITYNAME'])
    if col.value != None:
        #paste in ws2
        ws4.cell(row=s, column=headers_dic['PRIMARYSPECIALITYCODE'], value=p_spec[str(col.value).upper()])

# add secondary speciality if exists
for ss in range(2, ws4.max_row+1):
#     #copy from wb1
    col = ws4.cell(row=ss, column=headers_dic['SECANDRYSPECIALITYNAME'])
    pcol = ws4.cell(row=ss, column=headers_dic['PRIMARYSPECIALITYNAME'])
    if col.value != None and pcol.value != None:
        #check primary speciality code against secondry speciality code
        if p_spec[str(col.value).upper()] != p_spec[str(pcol.value).upper()]:
            ws4.cell(row=ss, column=headers_dic['SECANDRYSPECIALITYCODE'], value=p_spec[str(col.value).upper()])
# saving the destination excel file
wb4.save("D:\\middle.xlsx")

# Delete empty rows from middle
# xls = win32com.client.DispatchEx('Excel.Application')
# wb = xls.Workbooks.Open("D:\\middle.xlsx")
# ws = wb.Sheets('sheet')

# begrow = 1
# endrow = ws.UsedRange.Rows.Count
# for row in range(begrow,endrow+1): # just an example
#   if ws.Range('A{}'.format(row)).Value is None:
#     ws.Range('A{}'.format(row)).EntireRow.Delete(Shift=-4162) # shift up

# wb.Save()
# wb.Close()
# xls.Quit()
# ========================================= Validating ready sheet Data =================================================
for i in range (2, ws4.max_row + 1):
    for j in range (1, ws4.max_column + 1):
        # reading column header
        col_header = ws4.cell(row = 1, column = j)
        # reading cell value from source excel file
        cell = ws4.cell(row = i, column = j)
        if col_header.value == "NATIONALIDNUM":
            if cell.value == None or len(str(cell.value))<14:
                # write all row to incorrect sheet and skip rest of columns at same row
                ws3.cell(row = i, column = j).value = cell.value
                for y in range(1,ws4.max_column+1):
                    cc = ws4.cell(row = i,column = y)
                    ws3.cell(row = i, column = y).value = cc.value
                # add number of row to delete_row list
                del_rows.append(i)
            else:
                ws2.cell(row = i, column = j).value = cell.value
                ws2.cell(row = i, column = j).number_format = '0'
        elif col_header.value == "MOBILENUMBER":
        #     if cell.value == None or len(str(cell.value))<10:
        #         # write all row to incorrect sheet and skip rest of columns at same row
        #         ws3.cell(row = i, column = j).value = cell.value
        #         for y in range(1,ws4.max_column+1):
        #             cc = ws4.cell(row = i,column = y)
        #             ws3.cell(row = i, column = y).value = cc.value
        #         # add number of row to delete_row list
        #         del_rows.append(i)
        #     else:
            ws2.cell(row = i, column = j).value = str(cell.value)
            # ws2.cell(row = i, column = j).number_format = 'Text'
        #         # ws2.cell(row = i, column = j).alignment = center.align
        elif col_header.value == "PRACTITIONERNAMEENGLISH":
            if cell.value == None:
                # write all row to incorrect sheet and skip rest of columns at same row
                ws3.cell(row = i, column = j).value = cell.value
                for y in range(1,ws4.max_column+1):
                    cc = ws4.cell(row = i,column = y)
                    ws3.cell(row = i, column = y).value = cc.value
                # add number of row to delete_row list
                del_rows.append(i)
            else:
                #remove special characters
                result = ""
                for ch in str(cell.value):
                    # If char is not punctuation, add it to the result.
                    if ch not in string.punctuation:
                        result += ch
                        # writing the corrected value to ready excel file
                        ws2.cell(row = i, column = j).value = result
        elif col_header.value == "SHORTNAME":
            #remove special characters
            result = ""
            for ch in str(cell.value):
                # If char is not punctuation, add it to the result.
                if ch not in string.punctuation:
                    result += ch
                    # writing the read value to destination excel file
                    ws2.cell(row = i, column = j).value = result
        elif col_header.value == "GENDER":
            # writing the read value to ready excel file
            if cell.value == "Female":
                ws2.cell(row = i, column = j).value = "F"
            elif cell.value == "Male":
                ws2.cell(row = i, column = j).value = "M"
            else:
                ws2.cell(row = i, column = j).value = str(cell.value).upper()
        else:
            # writing the read value to ready excel file
            ws2.cell(row = i, column = j).value = cell.value

if del_rows:
    del_rows.sort()
    unique = set(del_rows)
    del_rows = list(unique)
    for item in del_rows:
        ws2.delete_rows(item - del_rows.index(item))
# #print(del_rows-1)
wb2.save("D:\\ready.xlsx")
# change columns name of ready file before saving
df = pd.read_excel("D:\\ready.xlsx")
df.rename(columns={'FACILITYID':'FACILITY_ID','PRACTITIONERTYPECODE':'PRACTITIONER_TYPE_CODE','PRACTITIONERID':'PRACTITIONER_ID','PRACTITIONERNAMEENGLISH':'PRACTITIONER_NAME_ENGLISH','PRACTITIONERNAMEARABIC':'PRACTITIONER_NAME_ARABIC','NATIONALIDNUM':'NATIONAL_ID_NUM','DATEOFBIRTH':'DATE_OF_BIRTH','MOBILENUMBER':'MOBILE_NUMBER','SECANDRYSPECIALITYCODE':'SECANDRY_SPECIALITY_CODE','MAINSTORE':'MAIN_STORE','OPSTORE':'OP_STORE','IPSTORE':'IP_STORE','ERSTORE':'ER_STORE','PRIMARYPHARMACIST':'PRIMARY_PHARMACIST','SHORTNAME':'SHORT_NAME','PRIMARYSPECIALITYCODE':'PRIMARY_SPECIALITY_CODE','OUTPATIENTDEPARTMENT':'OUTPATIENT','INPATIENTDEPARTMENT':'INPATIENT'}, inplace = True)
df=df[['FACILITY_ID','PRACTITIONER_TYPE_CODE','PRACTITIONER_ID','PRACTITIONER_NAME_ENGLISH','PRACTITIONER_NAME_ARABIC','NATIONAL_ID_NUM','GENDER','DATE_OF_BIRTH','MOBILE_NUMBER','SECANDRY_SPECIALITY_CODE','OUTPATIENT','INPATIENT','PRACTITIONER_CONSULTANT','PRACTITIONER_SPECIALIST','MAIN_STORE','OP_STORE','IP_STORE','ER_STORE','PHARMACIST','PRIMARY_PHARMACIST','SHORT_NAME','PRIMARY_SPECIALITY_CODE']]
df = df.sort_values(by='PRACTITIONER_TYPE_CODE')
df.to_excel("D:\\ready.xlsx")

# delete empty rows from incorrect sheet
# for row in ws3:
#     remove(ws3)
# Save incorrect file
wb3.save("D:\\incorrect.xlsx")
# incorr_file = asksaveasfilename(title="Choose Path for incorrect Sheet",filetypes=[("Excel Workbook", ".xlsx")],defaultextension=".xlsx")
# if incorr_file:
#     # save incorrect names in seperate sheet
#     wb3.save(incorr_file)
# else: # user cancel the file browser window
#     print("No Incorrect File Created")

print (headers_dic)
print (del_rows)
# print (p_spec)

